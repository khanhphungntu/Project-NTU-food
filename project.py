''' This project is a combination of Google Map API, pygame library and openpuxl library. \
To run this program, run data.py, open mergesort.py, binary_search.py and popupbox.py'''

# These are all libraries needed for this program
import pygame 

import math

import openpyxl

import requests

from popupbox import *

from time import sleep

from data import *

from html.parser import HTMLParser


# initialize pygame
pygame.init()

# set initial size of display
display_width = 1000
display_height = 500

gameDisplay = pygame.display.set_mode((display_width,display_height)) # set mode for display
pygame.display.set_caption('NTU EAT NOW') # name of the application

clock = pygame.time.Clock()

image = pygame.image.load("Mapgg.PNG") #load image of the map to image
surface_width=1300
surface_height=932
image = pygame.transform.scale(image,(surface_width,surface_height))



x_choose = -100 # variable determining the chosen location on the map
y_choose = -100 # variable determining the chosen location on the map

x = 0 # variable controling the left top corner position of image
scroll_y = 0 # variable controlling left top corner position of image

y=0

current_string=[] # is used to store data input from user
geo_location_x = 0 # variable to store value of geoposition on Google Map
geo_location_y = 0

list_canteen_sortedbydistance = [] # list of canteen sorted by distance


#calculate distance between chosen location and distance between canteen
def calculatedistance( y_choose,x_choose):
    global list_Canteen
    food_court_sorted = [] # food court sorted by distance (saved as object)
    distance_data = [] # distance calculated to each food court

    for i in list_Canteen: #calculating distance between choosen location and canteens
        destination_y = i.location_on_map[0]
        destination_x = i.location_on_map[1]
        r = requests.get("https://maps.googleapis.com/maps/api/directions/json?origin="+str(y_choose)+","+str(x_choose)+"&destination="+str(destination_y)+","+str(destination_x)+"&optimize_waypoints=True&key=AIzaSyC7YmRvn4WuGbFEG1Yy5DrQ3RM2VtuNXtU")
        distance = r.json()["routes"][0]["legs"][0]["distance"]["value"]
        distance_data.append(distance)
        food_court_sorted.append(i)
    #sorting the canteen from the nearest to the furthest using bubble sort
    for i in range(len(distance_data)-1):
        swap = 0
        
        for k in range(len(distance_data)-1-i):
            
            if distance_data[k] > distance_data [k+1]:
                
                distance_data[k] ,distance_data[k+1] = distance_data[k+1] ,distance_data[k]
                
                food_court_sorted[k] ,food_court_sorted[k+1] = food_court_sorted[k+1] ,food_court_sorted[k]
                
                swap = 1
        
        if swap == 0:
            break
    for i in range(len(distance_data)):
        distance_data[i] = str(distance_data[i])+ " m"
    return distance_data,food_court_sorted

i = 1 #zoom control variable

def zoom(): # zoom function

    global i, carImg, surface_width, surface_height, x, scroll_y,display_height, display_width
    mouse_x = get_mouse_position()[0]
    mouse_y = get_mouse_position()[1]
    if 790 < mouse_x < 801 and display_height - 12 < mouse_y < 500  and i<3:
        #zoom in the map 
        i+=1
    elif 820 < mouse_x < 831 and display_height - 12 < mouse_y < 500 and i>1:
        #zoom out the map
        i  -= 1 
        scroll_y = 0
        x = 0


def map_display(image, x,scroll_y): # fuction bliting image of map
    gameDisplay.blit(image, (x,scroll_y))

def get_mouse_position(): #get position of mouse
    mouse_position = pygame.mouse.get_pos()
    return mouse_position

def cover_page(): # displaying the cover page of application
    global gameDisplay, display_height, display_width,control
    crashed = False
    welcome_image = pygame.image.load("backpage.jpg") #load background image
    
    welcome_image = pygame.transform.smoothscale(welcome_image,(display_width,display_height))
   
    welcome_message = pygame.font.SysFont("Candara",50) #display welcome message
   
    start_button = pygame.font.SysFont("Book Antiqua",50) #display start button
   
    # display text
    welcome_image.blit(welcome_message.render("Welcome to NTU food", 1, (0,0,100)),(440,200))

    while not crashed:
        
        for event in pygame.event.get():
            gameDisplay.blit(welcome_image, (0,0))
            if event.type == pygame.QUIT: # if user want to quit the program
                pygame.quit()
                quit()
            
            if event.type == pygame.MOUSEBUTTONDOWN and event.button== 1:
                if 590< get_mouse_position()[0] < 700 and 270 < get_mouse_position()[1] < 330:

                    crashed = True # if user click start then continue
                    control = 0 # variable controling the flow of program
        
        if 590< pygame.mouse.get_pos()[0] < 700 and 270 < pygame.mouse.get_pos()[1] < 330:
            welcome_image.blit(start_button.render("Start",1, (0,0,100)),(590,270) ) # if mouse position in the start button the color of button changes
        
        else:
            welcome_image.blit(start_button.render("Start",1, (200,0,0)),(590,270) ) 
            # displaying the start button
       
        pygame.display.update()
        clock.tick(60) # settiing the number of framers per second

def move_map(): # move the map using four directions presses
    global x, scroll_y
    keys = pygame.key.get_pressed()
    
    if keys[pygame.K_LEFT]: # if user press left
        x = min( x + 30,0)
    
    elif keys[pygame.K_RIGHT]: # if user press right
        x = max( x-30, -surface_width*i + display_width )
    
    elif keys[pygame.K_DOWN]: # if user press down
        scroll_y = max(scroll_y - 30, -surface_height*i + display_height - 40)
    
    elif keys[pygame.K_UP]: # if user press up
        scroll_y = min(scroll_y + 30, 0)

def start_program(): # start the program 
    crashed = False # control variable to determine the condition to quit the start_program function
    global gameDisplay, list_canteen_sortedbydistance, calculatedistance,\
    quit_variable, x, scroll_y, current_string, control, list_canteen_sortedbydistance,\
    clock, x_choose, y_choose,surface_height, surface_width, i, display_width, display_height, geo_location_x, geo_location_y
    
    i = 1 # zoom control variable
    x = 0 # control variable of horizontal axis
    scroll_y = 0 #control variable of vertical axis
    # font to display
    text_display = pygame.font.SysFont("Calibri", 15)
    ask_to_choose_location = pygame.font.SysFont("Calibri", 18) 
    
    #load background image
    simple_background = pygame.image.load("simplebackground.jpg")
    simple_background= pygame.transform.smoothscale(simple_background,(1366,768))
    
    change_question = 0 # control variable to change text display
    while not crashed:
        #move the map in four directions
        move_map()
        for event in pygame.event.get():
            # load and display image

            if event.type == pygame.VIDEORESIZE: # controlling the size of window
                display_height = event.h
                display_width = event.w
                gameDisplay = pygame.display.set_mode((display_width,display_height),pygame.RESIZABLE)

            #get location of mouse or location of user on the map
            if event.type == pygame.MOUSEBUTTONDOWN and event.button== 1: # determing the chosen location
                
                if get_mouse_position()[1] <display_height -40: 
                    # get position of the mouse
                    x_choose = (get_mouse_position()[0] - x) 
                    y_choose = (get_mouse_position()[1] - scroll_y)
                    
                    geo_location_x = x_choose/surface_width/i*0.014163 + 103.676233 # calculate the geolocation 
                    geo_location_y = y_choose/surface_height/i*-0.016496 + 1.356310
                   
                    # request google api for selected location
                    r = requests.get("https://maps.googleapis.com/maps/api/geocode/json?latlng="+str(geo_location_y)+","+str(geo_location_x)+"&mode=walking&key=AIzaSyC7YmRvn4WuGbFEG1Yy5DrQ3RM2VtuNXtU")
                    location_real = r.json()["results"][0]["formatted_address"]
                    change_question = 1
                
                elif 520 < get_mouse_position()[0] < 670 and display_height - 30 < get_mouse_position()[1] < display_height -5: 
                    #condition to continue after choosing location
                    list_canteen_sortedbydistance = calculatedistance(geo_location_y, geo_location_x)[1] 
                    # calculating distance from chosen location
                    control = 1
                
                if display_width -221 < get_mouse_position()[0] < display_width - 199 and display_height - 23 < get_mouse_position()[1] < display_height - 1  and i<3:
                    #zoom in the map using left bracket
                    i +=1
                elif display_width - 191 < get_mouse_position()[0]< display_width - 169 and display_height - 23< get_mouse_position()[1] < display_height -1 and i>1:
                    #zoom out the map using right bracket
                    i  -= 1 
                    scroll_y = 0
                    x = 0
            if event.type == pygame.QUIT: # condition to quit the program
                pygame.quit()
                quit()
        
        #escape from the while loop and continue next function
        if control == 1:
            crashed = True
        # load and display image
        image = pygame.image.load("Mapgg.PNG")
        image = pygame.transform.scale(image, (surface_width*i,surface_height*i))
        map_display(image, x,scroll_y) #blit image
        
        gameDisplay.blit(simple_background,(0,display_height - 40)) # blit background to display text
        # blit text
        gameDisplay.blit(text_display.render("Move map by pressing four direction keys", 1, (0,0,0)),( display_width - 300, display_height - 40) )
        gameDisplay.blit(text_display.render("Zoom: ", 1, (0,0,0)),( display_width - 300, display_height - 20) )
        # ask user for selecting location
        if change_question == 0:
            gameDisplay.blit(ask_to_choose_location.render("Please choose your location:", 1, (0,0,0)),(5, display_height - 20) )
        
        # ask user if their position is accurate
        elif change_question == 1:
            gameDisplay.blit(ask_to_choose_location.render("Are you at %s? if true press continue" %(location_real), 1, (0,0,0)),(15, display_height - 25) )
            pygame.draw.rect(gameDisplay, (200,200,200),(520,display_height -30, 150,25), 0)
            pygame.draw.rect(gameDisplay, (0,0,0),(520,display_height -30, 150,25), 1)
            
            # display continue button
            continue_button = pygame.font.SysFont("Time New Roman", 25)
            if 488 < get_mouse_position()[0] <688 and display_height -30 < get_mouse_position()[1] < display_height- 5:
                gameDisplay.blit(continue_button.render("Continue",1,(100,100,0)),(540,display_height-25)) 
            else:
                gameDisplay.blit(continue_button.render("Continue",1,(0,100,0)),(540,display_height -25))
        # display button of zoom in and zoom out
        pygame.draw.circle(gameDisplay, (0,200,255),(display_width-210, display_height -12), 11,11 )
        pygame.draw.circle(gameDisplay, (0,200,255),(display_width - 180, display_height -12), 11,11 )
        pygame.draw.line(gameDisplay,(0,0,0),(display_width - 221,display_height -12),(display_width - 199, display_height -12))
        pygame.draw.line(gameDisplay,(0,0,0),(display_width - 210,display_height -23),(display_width - 210, display_height -1))
        pygame.draw.line(gameDisplay,(0,0,0),(display_width - 191,display_height -12),(display_width - 169, display_height -12))      
        
        pygame.display.update()# update to the display
        clock.tick(60) # set fps

# load image of button
frame = pygame.image.load("button.png")
frame = pygame.transform.smoothscale(frame, (180,35))

def back(): # back button
    global  gameDisplay,frame
    back_to_map_text = pygame.font.SysFont("Candara", 25)
    gameDisplay.blit(frame, (5,5))
    gameDisplay.blit(back_to_map_text.render("Back",1,(0,0,0)) ,(65,10))


def ask_criteria(): # this function aims to ask user for choosing their criteria
    crashed = False
    global gameDisplay, back, conditon_to_back, user_input, quit_variable, x_choose, y_choose, carImg, x, scroll_y, current_string, control, clock, surface_height, surface_width, i, display_width, display_height
    
    simple_background = pygame.image.load("dark.jpg") #load background image
    simple_background= pygame.transform.smoothscale(simple_background,(1366,900))
    # load anddisplay the frame of button and long bar
    long_bar = pygame.image.load("longbar.png")
    long_bar = pygame.transform.smoothscale(long_bar, (600, 65))
    button = pygame.image.load("button.png")
    button = pygame.transform.smoothscale(button, (200, 80))
    button_1 = pygame.transform.smoothscale(button, (300, 80))
    
    # text font to display
    text_1 = pygame.font.SysFont("Candara", 30 )
    text_2 = pygame.font.SysFont("Book Antiqua", 25 )
    
    while not crashed:
        # two variables to store mouse positions
        x_mouse = get_mouse_position()[0]
        y_mouse = get_mouse_position()[1]
        for event in pygame.event.get():
            if event.type == pygame.VIDEORESIZE: #resize window
                display_height = event.h
                display_width = event.w
                gameDisplay = pygame.display.set_mode((display_width,display_height),pygame.RESIZABLE)
            
            if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1: 
                # if user press back button
                if 5 < x_mouse <205 and 5 < get_mouse_position()[1] <35:
                    control -= 1
                    crashed = True
                #if user choose button distance criteria
                elif 120< x_mouse <320 and 200< y_mouse < 280 : 
                    control = 2
                    crashed = True

                # if user choose food criteria
                elif 120 +800/3< x_mouse< 320 + 800/3 and 200 < y_mouse< 280:
                    control = 3
                    crashed = True
                # if user chooses type of cuisine
                elif 120 + 800/3*2< x_mouse< 320 + 800/3*2 and 200 < y_mouse< 280: 
                    control = 6
                    crashed = True
                # if user choose update information button
                elif 300 < y_mouse < 380 and 350 < x_mouse < 650:
                    control = 5
                    crashed = True

            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
        #display new background
        gameDisplay.blit(simple_background,(0,0))
        #display the question
        gameDisplay.blit(long_bar, (200, 100 ))
        text_main = text_1.render("What is your Criteria?", 1, (0,50,100))
        text_main_rect = text_main.get_rect( center = (500,100+65/2))
        gameDisplay.blit(text_main, text_main_rect)
        
        # display the text in the button
        list_text = ["Distance", "Food", "Type of Cuisine"] # list of button text to display
        
        for i in range(3): # display text to button
            gameDisplay.blit(button, (120 + 800/3*i, 200 ))
            text_to_display = text_2.render("%s"%(list_text[i]), 1, (0,0,100))
            text_rect = text_to_display.get_rect( center = (220 + 800/3*i, 240))
            gameDisplay.blit(text_to_display, text_rect )
        # display the requirement to update information
        gameDisplay.blit(button_1, (350,300))
        text_update = text_2.render("Update information",1,(0,0,100))
        text_update_rect = text_update.get_rect(center = (500,340))
        gameDisplay.blit(text_update, text_update_rect)
        
        back() # back to choosing location

        pygame.display.update() # update to display
        clock.tick(60) # set fps

def distance_criteria(): # if user choose distance is the forefront criteria
    global x_choose, control, quit_variable,\
     y_choose, back, conditon_to_back, gameDisplay, \
     get_mouse_position, display_width,display_height, control_distance_case, geo_location_y, geo_location_x, list_canteen_sortedbydistance
    
    crashed = False
    white_background = pygame.image.load("whitecover.jpg") #load background image
    while not crashed:
        
        for event in pygame.event.get():
            # get the position of mouse
            mouse_x = get_mouse_position()[0]
            mouse_y = get_mouse_position()[1]
            if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
                if 100 <mouse_x < 500 and 100 < mouse_y < 610:
                    control_distance_case = 1 # start controling the next program by this variable
                    # sorted canteen by distance calculated from google map
                    try:
                        information_of_chosen_canteen(list_canteen_sortedbydistance[int((mouse_y-100)/34.5)]) 
                    except Exception:
                        pass

                if 5 < mouse_x <205 and 5 < mouse_y <35: # click back button
                    control -= 1
                    crashed = True
            
            if event.type == pygame.VIDEORESIZE: # resize window 
                display_height = event.h
                display_width = event.w
                gameDisplay = pygame.display.set_mode((display_width,display_height),pygame.RESIZABLE)
            
            if event.type == pygame.QUIT: # quit program
                pygame.quit()
                quit()
        
        white_background = pygame.transform.smoothscale(white_background, (display_width,display_height))
        gameDisplay.blit(white_background, (0,0)) #display background 
        back() # back to choosing location
        # display text to the screen
        list_of_canteen_text = pygame.font.SysFont("Book Antiqua", 45)
        gameDisplay.blit(list_of_canteen_text.render("List of canteens sorted by distance:",1,(0,0,50)), (200,35))
        name_canteen_text_display = pygame.font.SysFont("Candara", 27)
        
        #display table of canteen sorted by distance
        for i in range(12):
            pygame.draw.line(gameDisplay, (0,0,0), (100,100 + 34.5*i), (500, 100+ 34.5*i) )
            gameDisplay.blit(name_canteen_text_display.render("%s"%(list_canteen_sortedbydistance[i].name),1,(10,10,10)), (230, 110 +34.5*i))
            try:
                if 100 <mouse_x < 500 and 100 + 34.5*i < mouse_y < 134.5 + 34.5*i:
                    gameDisplay.blit(name_canteen_text_display.render("%s"%(list_canteen_sortedbydistance[i].name),1,(0,100,100)), (230, 110 +34.5*i))
                    popoup_info(list_canteen_sortedbydistance[i],int((mouse_y-100)/34.5),20, 35 )
            except UnboundLocalError:
                continue
        
        pygame.display.update() # update to the screen
        clock.tick(60) # set fps


def information_of_chosen_canteen(canteen_name): # display information about chosen canteen, this function is executed after user chooses canteen name
    global gameDisplay, list_Canteen, \
    quit_variable, display_height, display_width, information_of_chosen_stall
    
    stall_in_this_canteen = canteen_name.stalls # get all stalls in this canteen

    name_of_stall_in_this_canteen = [i.name for i in stall_in_this_canteen]
    crashed = False
    # display the image of star 
    star_tranparent = pygame.image.load("star_transparent.jpg")
    star_color = pygame.image.load("star.jpg")
    star_tranparent = pygame.transform.smoothscale(star_tranparent, (40,40))
    star_color = pygame.transform.smoothscale(star_color, (40,40))
    food_image = pygame.image.load("whitecover.jpg")
    number_of_stalls = len(name_of_stall_in_this_canteen)
    height_of_row = 415/number_of_stalls
    current_string1 = [] # store the characters input for min price 
    current_string2 = [] # store the characters input for max price
    # control variable
    ask_value = 0
    blit_star = 0
    number_of_star = 0
    text = pygame.font.SysFont("Arial",40) # text font
    please_choose_stall_text = pygame.font.SysFont("Candara", 40)
    text1 = pygame.font.SysFont("Book Antiqua", 30)
    while not crashed:
        gameDisplay.blit(food_image, (0,0))
        #display text and input box of price range
        pygame.draw.rect(gameDisplay, (0,0,0),(100, 200, 100, 50), 2)
        gameDisplay.blit(text.render("Min", 1, (0,0,50)), (30,200))
        pygame.draw.rect(gameDisplay, (0,0,0),(100, 280, 100, 50), 2)
        gameDisplay.blit(text.render("Max", 1, (0,0,50)), (30,280))
        # get mouse position
        mouse_x = get_mouse_position()[0]
        mouse_y = get_mouse_position()[1]        
        
        for event in pygame.event.get():
            
            if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
                
                if 300 < mouse_x < 700 and 80 < mouse_y < 495:
                    
                    try:
                        min_price = float("".join(current_string1[:3])) # get value from user input
                    except Exception:
                        min_price = 0
                    try:
                        max_price = float("".join(current_string2[:3])) # get value from user input
                    except Exception:
                        max_price = float("inf")
                    # execute function after user choose stall name
                    information_of_chosen_stall(canteen_name,stall_in_this_canteen[int((mouse_y-80)/height_of_row)], \
                                                min_price, max_price, number_of_star)
                
                elif 5 < mouse_x <205 and 5 < mouse_y <35:
                    crashed = True # back button

                elif 100 < mouse_x < 200 and 200 < mouse_y <250:
                    ask_value = 1  # user click to input min price

                elif 100 < mouse_x < 200 and 280 < mouse_y < 380:
                    ask_value = 2 #user click to input max price

                elif 0 <= (mouse_x-20)/40 <= 5 and 410 < mouse_y < 500  and event.type == pygame.MOUSEBUTTONDOWN:
                    blit_star = 1 # user click to input rank
                    number_of_star = (mouse_x-20)//40
            
            if event.type == pygame.VIDEORESIZE: # resize window
                display_height = event.h
                display_width = event.w
                gameDisplay = pygame.display.set_mode((display_width,display_height),pygame.RESIZABLE)
            
            if event.type == pygame.QUIT:# user click quit 
                pygame.quit()
                quit()

        food_image = pygame.transform.smoothscale(food_image, (display_width,display_height))

        gameDisplay.blit(please_choose_stall_text.render("Choose your foodstall", 1, (0,100,0)), (350,22))
        
        pygame.draw.rect(gameDisplay, (0,0,0),(300, 80, 400, 415), 1)
        
        name_of_stall_text = pygame.font.SysFont("Calibri", min(int(height_of_row) - 6, 25))
        
        for i in range(number_of_stalls): # effect if user move mouse through each item
            pygame.draw.line(gameDisplay, (0,0,0), (300, 80+ height_of_row*(i+1)), (700, 80+ height_of_row*(i+1) ))
            gameDisplay.blit(name_of_stall_text.render("%s .%s"%(i,name_of_stall_in_this_canteen[i]),1,(10,10,10)), (340, 83 + height_of_row*i))
            try:
                if 300 < mouse_x < 700 and 85 + height_of_row*i < mouse_y < 85 + height_of_row*(i+1):
                    gameDisplay.blit(name_of_stall_text.render("%s. %s"%(i,name_of_stall_in_this_canteen[i]),1,(100,0,100)), (340, 83 + height_of_row*i))
            except UnboundLocalError:
                continue

        if blit_star == 1: # blit the number of stars that user chooses for rank
            for k in range(number_of_star + 1):
                gameDisplay.blit(star_color, (20 + 40*k, 450))
            for j in range( 4- number_of_star):
                gameDisplay.blit(star_tranparent, (180 - 40*j, 450))
        else:
            for j in range(5):
                gameDisplay.blit(star_tranparent, (20 + 40*j, 450))
        
        back() # back function
        # display text to screen
        gameDisplay.blit(text1.render("Rank of food",1,(0,0,100)),(30,390))
        gameDisplay.blit(text1.render("Price",1,(0,0,100)),(60,150))
        display_box(gameDisplay, "".join(current_string1[:3]), 110, 220)
        display_box(gameDisplay, "".join(current_string2[:3]), 110, 300)
        
        if ask_value == 1:
            ask( event, current_string1) # get characters that user input to the min price
        elif ask_value == 2:
            ask(event, current_string2) # get characters that user input to the max price
        
        pygame.display.update()# update to display
        clock.tick(60) # set fps



def information_of_chosen_stall(canteen_name, stall_name, min_price, max_price, number_of_stars): # function executes after user input min price, max price ,rank and name of stall to the screen
    global gameDisplay, list_Canteen, search_food_by_price, \
    quit_variable, display_height, display_width, control_distance_case, information_of_chosen_stall, geo_location_x, geo_location_y
    crashed = False
    #this list contain all food satisfying the price range in the given stall
    list_of_satisfy_food = search_food_by_price((min_price, max_price), stall_name, number_of_stars)[:44]
    number_of_dishes = len(list_of_satisfy_food)
    # text font
    text_display = pygame.font.SysFont("Candara", 22)
    extra_information = pygame.font.SysFont("Arial", 15)
    please_choose_food_text = pygame.font.SysFont("Cambria", 30)
    please_choose_food_text_1 = pygame.font.SysFont("Cambria", 20)
    # control variable 
    sort_price = 0 
    get_direction = 0
    while not crashed:
        stall_image = pygame.image.load("whitecover.jpg") # load and display image
        stall_image = pygame.transform.smoothscale(stall_image, (display_width,display_height))
        gameDisplay.blit(stall_image, (0,0))

        for event in pygame.event.get():
            # get mouse position
            mouse_x = get_mouse_position()[0]
            mouse_y = get_mouse_position()[1]
            
            if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
                if 5 < mouse_x <205 and 5 < mouse_y <35: # back button
                    crashed = True
                elif 800 < mouse_x < 830 and 180 < mouse_y < 210 : # if user want to sort food by price
                    sort_price = 1
                    for i in range(len(list_of_satisfy_food)): # use bubble sort to sort food by price
                        swap = 0
                        for k in range(len(list_of_satisfy_food)-1-i):
                            if float(list_of_satisfy_food[k][1]) > float(list_of_satisfy_food[k+1][1]):
                                list_of_satisfy_food[k], list_of_satisfy_food[k+1] = list_of_satisfy_food[k+1], list_of_satisfy_food[k]
                                swap = 1
                        if swap == 0 :
                            break
            if event.type == pygame.VIDEORESIZE: # resize window
                display_height = event.h
                display_width = event.w
                gameDisplay = pygame.display.set_mode((display_width,display_height),pygame.RESIZABLE)
            
            if event.type == pygame.QUIT: # if user wants to quit the program
                pygame.quit()
                quit()
        # this part is used to display information of dishes to the screen based on the number of dishes
        if 0 == number_of_dishes:
            no_result = pygame.font.SysFont("Candara", 80)
            gameDisplay.blit(no_result.render("No result",1,(0,0,0)),(350,200))

        elif 1 <= number_of_dishes < 22:
            
            height_of_row = 445/number_of_dishes
            name_of_food_text = pygame.font.SysFont("Calibri", 18)
            for i in range(number_of_dishes):
                pygame.draw.line(gameDisplay, (0,0,0), (50, 50+ height_of_row*(i+1)), (530, 50+ height_of_row*(i+1) ))
                
                gameDisplay.blit(name_of_food_text.render("%s"%(list_of_satisfy_food[i][0][:40]),1,(10,10,10)), (70, 50 + height_of_row/3 + height_of_row*i))
                
                gameDisplay.blit(name_of_food_text.render("%s"%(list_of_satisfy_food[i][1]),1,(10,10,10)), (500, 50 + height_of_row/3 + height_of_row*i))
        
        elif  22<= number_of_dishes :
            
            height_of_row_table_1 = 445/22
            name_of_food_text = pygame.font.SysFont("Calibri", min(int(height_of_row_table_1) - 6, 25))

            for i in range(22):
                pygame.draw.line(gameDisplay, (0,0,0), (10, 50+ height_of_row_table_1*(i+1)), (310, 50+ height_of_row_table_1*(i+1) ))
                gameDisplay.blit(name_of_food_text.render("%s"%(list_of_satisfy_food[i][0][:40]),1,(10,10,10)), (15, 53 + height_of_row_table_1*i))
                gameDisplay.blit(name_of_food_text.render("%s"%(list_of_satisfy_food[i][1]),1,(10,10,10)), (290, 53+ height_of_row_table_1*i))
            
            remaining_dishes = min(number_of_dishes,44) - 22

            name_of_food_text = pygame.font.SysFont("Calibri", min(int(height_of_row_table_1) - 6, 25))
            
            # display food name
            for i in range(remaining_dishes-1):
                pygame.draw.line(gameDisplay, (0,0,0), (350, 50+ height_of_row_table_1*(i+1)), (660, 50+ height_of_row_table_1*(i+1) ))
                gameDisplay.blit(name_of_food_text.render("%s"%(list_of_satisfy_food[i+23][0][:40]),1,(10,10,10)), (370, 53 + height_of_row_table_1*i))
                gameDisplay.blit(name_of_food_text.render("%s"%(list_of_satisfy_food[i+23][1]),1,(10,10,10)), (640, 53 + height_of_row_table_1*i))
        
        pygame.draw.rect(gameDisplay,(0,0,0),(800,180, 30,30), 1)
        # if user wantto sort food by price
        if sort_price == 1:
            pygame.draw.rect(gameDisplay,(0,0,100),(800,180, 30,30), 0)
        gameDisplay.blit(text_display.render("Sort by Price", 1, (0,50,0)), (670,185))
        # display extra information about canteen
        if len(str(stall_name.name)+ str(canteen_name.name)) > 30:
            gameDisplay.blit(please_choose_food_text_1.render("%s in %s"%(stall_name.name, canteen_name.name), 1, (0,100,0)), (300,15))
        else:
            gameDisplay.blit(please_choose_food_text.render("%s in %s"%(stall_name.name, canteen_name.name), 1, (0,100,0)), (300,15))
        gameDisplay.blit(extra_information.render("Address: %s %s"%(canteen_name.name, canteen_name.real_location),1,(0,0,0)),(680,230))
        gameDisplay.blit(extra_information.render(" Capacity: %s "%(canteen_name.seat_capacity),1,(0,0,0)),(680,260))
        gameDisplay.blit(extra_information.render(" Time open:%s Phone: %s"%(canteen_name.time_open, canteen_name.phone),1,(0,0,0)),(680,290))
        # back function
        back()
        # if user want to get direction to the canteen, function get direction will be implemented
        try:
            get_direction = button_get_direction(event, mouse_x, mouse_y)   
        except Exception:
            pass
        if get_direction == True:
            pop_up_direction(geo_location_y, geo_location_x, canteen_name)
        # update to display 
        pygame.display.update()
        clock.tick(60) # set fps


def search_food_by_enter_name(): # this function allow user to choose food by entering name
    global display_height, display_width, back,get_mouse_position, gameDisplay, control, geo_location_x, geo_location_y
    current_string = [] # this list is to stored the characters input by users
    
    crashed = False
    start_to_type = 0 
    name_of_wanted_food = "" # this varibale contain the food name of user's choice
    button_color_price = 0 # control user's choice with price
    button_color_rank = 0 # control user's choice with rank
    #font to display
    text_display = pygame.font.SysFont("Calibri", 33)
    fontobject = pygame.font.SysFont("Candara",25)
    sort_text = pygame.font.SysFont("Candara", 25)
    information_text = pygame.font.SysFont("Arial", 20)
    small_text= pygame.font.SysFont("Calibri", 18)
    text = pygame.font.SysFont("Calibri",25)
    text_display_1 = pygame.font.SysFont("Calibri", 20)
    # control variable
    canteen_to_get_detail = -1
    get_detail = 0
    pop_up = 0
    get_direction = 0
    while not crashed:
        # get mosue position
        mouse_x = get_mouse_position()[0]
        mouse_y = get_mouse_position()[1]
        gameDisplay.fill((250,250,250)) # fill the screen with white color
        for event in pygame.event.get():
            if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
                if canteen_to_get_detail > -1 and 800 <mouse_x < 957 and 470 < mouse_y < 497 :
                    pop_up  = 1
                elif 30 < mouse_x <430 and 98 < mouse_y < 169:
                    start_to_type = 1 # user clicks the search bar 
                elif 278 < mouse_x < 312 and 58 < mouse_y < 92:
                    button_color_price = 1 # user chooses to sort data by price
                    button_color_rank = 0
                elif 395 < mouse_x < 435 and 58 < mouse_y < 92:
                    button_color_price = 0
                    button_color_rank = 1# user choose to sort data by name
                elif 5< mouse_x < 205 and 5 < mouse_y < 35:
                    crashed = True # back button
                    control = 1
                elif 90 < mouse_y < 470 and 512 < mouse_x < 955 and get_detail == 1:
                    canteen_to_get_detail = (mouse_y - 90)*len(information)//400 # if user clicked on a suggested result on the screen
                
                elif pop_up == 1 and 484 < mouse_x <  517 and 443 < mouse_y < 476:
                    pop_up = 0 
            if event.type == pygame.VIDEORESIZE: # resize window
                display_height = event.h
                display_width = event.w
                gameDisplay = pygame.display.set_mode((display_width,display_height),pygame.RESIZABLE)
            elif event.type == pygame.QUIT:
                pygame.quit()
                quit()
        pygame.draw.line(gameDisplay, (0,0,0), (500,50),(500,450))
        gameDisplay.blit(text_display.render("Search food by name",1, (0,0,50)),(400,15))
        pygame.draw.rect(gameDisplay, (0,0,0), (30,98,450,50),1)
        if start_to_type == 0:
            gameDisplay.blit(text.render("Type your food here",1,(0,0,0)),(40,108))
        
        if button_color_price == 1:
            pygame.draw.rect(gameDisplay, (0,0,100), (280,60,30,30),0)
        else:
            pygame.draw.rect(gameDisplay, (0,0,0), (280,60,30,30),1)
        if button_color_rank == 1:
            pygame.draw.rect(gameDisplay, (0,0,100), (400,60,30,30),0)
        else:
            pygame.draw.rect(gameDisplay, (0,0,0), (400,60,30,30),1)
        
        # text display on the screen
        gameDisplay.blit(sort_text.render("Sort by : Price",1, (10,10,100)),(120,65))
        gameDisplay.blit(sort_text.render("Rank",1, (10,10,100)),(335,65))
        gameDisplay.blit(text_display_1.render("Location",1, (0,50,100)),(580,50))
        gameDisplay.blit(text_display_1.render("Price",1, (0,50,100)),(800,50))
        gameDisplay.blit(text_display_1.render("Rank",1, (0,50,100)),(890,50))
        gameDisplay.blit(small_text.render("(Choose the stall and get detail)",1, (0,0,50)),(520,65))
        result_of_search = search_foodname_by_characters("".join(current_string)[:12])
        
        if start_to_type == 1:
            ask(event, current_string) # this function is used to store the data input bu user
            if len(current_string) != 0: # display to the screen the input of user
                gameDisplay.blit(fontobject.render("".join(current_string)[:40], 1, (0,0,50)),(50,110))
            
            # this part is to show related result to the input
            number_of_results = len(result_of_search)
            if number_of_results != 0:
                for i in range(number_of_results):
                    pygame.draw.line(gameDisplay, (0,0,0), (50,193+ 33*i),(450,193+ 33*i))
                    if 50 < mouse_x < 450 and 160 < mouse_y < 160 + 33*number_of_results and i == int((mouse_y-160)/33):
                        if event.type == pygame.MOUSEBUTTONDOWN and event.button ==1: # if user clicked on the suggested food on the screen
                            name_of_wanted_food = result_of_search[i] # get the name of chosen food
                            current_string = [i for i in name_of_wanted_food] # display the chosen suggested food to search bar
                        gameDisplay.blit(text.render("%s"%(result_of_search[i][:40]),1,(0,0,200)), (80,165+ 330/10*i))
                    else:
                        gameDisplay.blit(text.render("%s"%(result_of_search[i][:40]),1,(0,0,0)), (80,165+ 330/10*i))
          
        # this part is to provide the information related to chosen food_name
        
        if name_of_wanted_food != "":
            # search food after user clicks on the suggested food name on the screen
            get_detail = 1
            information = search_food_by_name(name_of_wanted_food, button_color_price, button_color_rank) # search food by name and order them by rank or price
            for i in range(len(information)):
                pygame.draw.line(gameDisplay, (0,0,0), (512, 90 + 400/len(information)*i),(962, 90+ 400/len(information)*i))
                # display the result of search to the screen with related information
                if i == canteen_to_get_detail:
                    gameDisplay.blit(information_text.render("%s"%(information[i][2]),1, (50,200,200)),(820,95 + 400/len(information)*i))
                    gameDisplay.blit(information_text.render("%s"%(information[i][3]),1, (50,200,200)),(890,95 + 400/len(information)*i))
                    location_display = str(information[i][1]+ information[i][0].name)[:30]
                    gameDisplay.blit(information_text.render("%s"%(location_display),1, (50,200,200)),(518,95 + 400/len(information)*i))
                else:
                    gameDisplay.blit(information_text.render("%s"%(information[i][2]),1, (0,50,100)),(820,95 + 400/len(information)*i))
                    gameDisplay.blit(information_text.render("%s"%(information[i][3]),1, (0,50,100)),(890,95 + 400/len(information)*i))
                    location_display = str(information[i][1]+ information[i][0].name)[:30]
                    gameDisplay.blit(information_text.render("%s"%(location_display),1, (0,50,100)),(518,95 + 400/len(information)*i))                    
        
        if canteen_to_get_detail >= 0 : # effects of changing color of detail button after user clicks canteen
            pygame.draw.rect(gameDisplay, (200,200,200), (800,470,150,27),0)
            pygame.draw.rect(gameDisplay, (0,0,0), (800,470,150,27),1)
            gameDisplay.blit(information_text.render("Get detail",1,(0,0,100)),(830,473))
        else:
            pygame.draw.rect(gameDisplay, (0,0,0), (800,470,150,27),1)
            gameDisplay.blit(information_text.render("Get detail",1,(0,0,0)),(830,473))
        if pop_up == 1: # user clicks detail button to get more information about the canteen
            try:
                pop_up_box(information[canteen_to_get_detail],event, name_of_wanted_food)
            except Exception:
                pass
        
        back() # back button
        try: # get direction button to guide user the route to get to the canteen
            get_direction = button_get_direction(event, mouse_x, mouse_y)
        except Exception:
            pass
        if canteen_to_get_detail >=0 and get_direction == True:
            try:
                pop_up_direction(geo_location_y, geo_location_x, information[canteen_to_get_detail][0]) # get direction function is implemented if user clicks get direction button
            except Exception:
                pass
        pygame.display.update() # update to the screen
        clock.tick(60) # set fps

def pop_up_box(information,event, name_food): # this function is executed if user click get detail button on the screen
    global gameDisplay
    # load and display pop_up frame
    pop_up_image = pygame.image.load("popupnow.png")
    pop_up_image = pygame.transform.smoothscale(pop_up_image, (600,200))
    gameDisplay.blit(pop_up_image, (200, 290))
    # text font
    text_1 = pygame.font.SysFont("Book Antiqua", 20)
    text_2 = pygame.font.SysFont("Book Antiqua", 15)
    # display information to the pop up frame
    gameDisplay.blit(text_1.render("%s in %s"%(name_food, information[1]),1,(0,0,100)), (215, 300))
    gameDisplay.blit(text_2.render("Address : %s, %s "%(information[0].name, information[0].real_location),1,(0,0,100)), (215, 340))
    gameDisplay.blit(text_2.render("Time open: %s  Phone %s"%(information[0].time_open,information[0].phone),1,(0,0,100)), (215, 380))
    gameDisplay.blit(text_2.render("Seat Capacity: %s"%(information[0].seat_capacity),1,(0,0,100)), (215, 420))


def update_information(): # this function allows user to update information such as food, price, rank of food to the list of food
    global gameDisplay, list_Canteen, update_now, control, display_width, display_height
    background =pygame.image.load("light.jpg")
    crashed = False
    # list of text displayed to the screen
    list_of_word = ["Canteen name", "Stall name", "Food name", "New price", "New rank"]
    start_to_type = 0
    #string store data to be updated, from left to right is list of characters of canteen name, stall name, food name, price, rank
    string = [[],[],[],[],[]]
    # all font to display
    text_display = pygame.font.SysFont("Candara", 40)
    text_display_1 = pygame.font.SysFont("Arial", 25)
    text_display_2 = pygame.font.SysFont("Calibri", 15)
    text_display_3 = pygame.font.SysFont("Calibri", 25)
    text_display_4 = pygame.font.SysFont("Arial", 30)
    # data to display stores the data for the next display if user choose input 
    data_to_display = []
    # control variable of function
    next_data = ""
    result =""
    start_value= 0
    while not crashed:
        # load and display image to the display
        background = pygame.transform.smoothscale(background, (display_width, display_height))
        gameDisplay.blit(background, (0,0))
        # get mouse position
        mouse_x = get_mouse_position()[0]
        mouse_y = get_mouse_position()[1]
        for event in pygame.event.get():
            if event.type == pygame.VIDEORESIZE: # resize window
                display_height = event.h
                display_width = event.w
                gameDisplay = pygame.display.set_mode((display_width,display_height),pygame.RESIZABLE)
            elif event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
                if result == True or result == False:
                    result = "" # if user do not click update button
                if 5 < mouse_x <205 and 5 < mouse_y <35: # back button
                    crashed = True
                    control = 1
                # recognize if user is choosing one inputbox
                elif 150 < mouse_x < 500 and 0 < (mouse_y-95)/65 < 5:
                    start_to_type = 1
                    number_position = 1
                    position_to_input = (mouse_y-95)//65
                    start_value = 0
                # if user choose to update data button
                elif 110 < mouse_x < 260 and 430 < mouse_y < 470:
                    result = update_now(string)
            elif event.type == pygame.QUIT:
                pygame.quit()
                quit()
        gameDisplay.blit(text_display.render("Update Information",1, (0,0,100)), (350,20))
        # recognize that if user choose to start to type
        if start_to_type == 1:
            # store characters that user input to the screen
            ask(event, string[position_to_input])
            if position_to_input == 0:
                
                data_to_display = [i for i in list_Canteen]
            elif position_to_input == 1 and next_data in list_Canteen:
                data_to_display = next_data.stalls
            elif position_to_input == 2 and type(next_data) == Stall  :
                data_to_display = next_data.menu
        gameDisplay.blit(text_display_1.render("Type Here", 1, (0,0,0)), (155, 50)) # display text to the screen
        # display the input information to the search bar
        for i in range(5):
            pygame.draw.rect(gameDisplay, (0,0,0), (150, 95 + 65*i, 350, 40), 1)
            gameDisplay.blit(text_display_1.render("%s"%(list_of_word[i]), 1, (0,0,0)), (15,100 + 65 *i))   
            if len(string[i]) > 27 :
                gameDisplay.blit(text_display_2.render("".join(string[i]), 1, (0,0,0)),(158, 110 + 65 *i))
            else:
                gameDisplay.blit(text_display_3.render("".join(string[i][:27]), 1, (0,0,0)),(158, 100 + 65 *i))
        gameDisplay.blit(text_display_2.render("Rank is 1 to 5", 1, (0,100,0)),(160, 400)) # display text to the screen
        # effect if user click update button
        if 110 < mouse_x < 260 and 430 < mouse_y < 470:
            pygame.draw.rect(gameDisplay, (200,200,200), (110, 430, 150, 40), 0)
            pygame.draw.rect(gameDisplay, (0,0,0), (110, 430, 150, 40), 1)
            gameDisplay.blit(text_display_4.render("Update",1, (0,100,100)), (135, 430))
        else:
            pygame.draw.rect(gameDisplay, (0,0,0), (110, 430, 150, 40), 1)
            gameDisplay.blit(text_display_4.render("Update",1, (0,100,100)), (135, 430))
        
        # this part is for user select the suggested result 
        if start_to_type == 1:
            (next_data, start_value, number_position) = page(data_to_display, event, next_data, position_to_input, mouse_x, mouse_y,string, start_value,number_position)
        # check if the data input is suitable to be displayed to the screen
        if result == True :
            gameDisplay.blit(text_display_1.render("Done!", 1, (100,100,0)), (280, 430))
        elif result == False:
            gameDisplay.blit(text_display_1.render("Type Error", 1, (100,100,0)), (280, 430))
        
        back() # back button
        pygame.display.update() # update information to the screen
        clock.tick(60) # set fps

# this function is to display the data in pages if the number of data is too large
def page(data_to_display, event, next_data, position_to_input, mouse_x, mouse_y,string,start_value, number_position):
    # set text font
    text_display_5 = pygame.font.SysFont("Calibri", 19)
    text_display_6 = pygame.font.SysFont("Arial Rounded MT", 22)
    try: # if user choose to input name of canteen and name of food stall
        if type(data_to_display[0]) == type((1,2)):
            data_string = [i[0] for i in data_to_display]
        else:# if user chooses to input name of food
            data_string = [i.name for i in data_to_display]
    except Exception:
        pass
    
    for i in range(4):# display the number of page of information
        gameDisplay.blit(text_display_6.render("%s"%(i+1), 1, (0,0,0)), (650 + i*30, 60 ))
    # get the position of user input 
    if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
        if 645 < mouse_x < 755 and 58< mouse_y < 70:
            number_position = (mouse_x - 640)//30 +1
            start_value = number_position - 1
        elif 580 < mouse_x < 990 and 80 < mouse_y < 480 and position_to_input < 3:
            blit_position = int((mouse_y - 80)/400*15)
            string[position_to_input] = list(data_string[blit_position + 15*start_value])
            if type(data_to_display[0]) == type((1,2)):
                next_data = []
            else:
                next_data = data_to_display[blit_position + 15*start_value]
    # change the color of page information
    gameDisplay.blit(text_display_6.render("%s"%(number_position), 1, (200,0,200)), (650 + (number_position-1)*30, 60 ))
    # display suggested output to the screen
    for i in range(15):
        pygame.draw.line(gameDisplay, (0,0,0), (580, 100 + 400/15*i),(880, 100 + 400/15*i ))
        try:
            gameDisplay.blit(text_display_5.render("%s"%(data_string[i + 15*start_value]), 1, (0,0,0)), (590, 83 + 400/15*i ))
        except Exception:
            continue
    return next_data, start_value, number_position


# this function is to check the information get from user input and update the data into excel file
# if the name of canteen, name of stall, name of food are true data is updated for price and rank
# elif the name of canteen, name of stall is true but food name is not in stall, this will be updated as new item
# if food stall or canteen name is false, type error
# all data will be updated to excel file for future use
def update_now(string):
    global list_Canteen
    book = openpyxl.load_workbook('Canteen-Copy.xlsx') # open excel file to update data
    canteen_name = "".join(string[0])
    stall_name = "".join(string[1])
    food_name = "".join(string[2])
    new_price = "".join(string[3])
    new_rank = "".join(string[4])
    try:
        float(new_price) # check if price is correct value
        int(new_rank) # check if rank is correct value
    except Exception:
        return False

    if int(new_rank) > 5 or int(new_rank) <0 : # rank is from 1 to 5 if out of this range update function is false
         return False
    for i in range(len(list_Canteen)): # search for the canteen name as input by user, if no result, return false
        if check_character(canteen_name, list_Canteen[i].name):
            canteen_name = list_Canteen[i]
            sheet = book.get_sheet_by_name(list_Canteen[i].name)
            break
    else:
        return False
    for k in range(len(canteen_name.stalls)): # search for the stall name as input by user, if no result, return false
        if check_character(canteen_name.stalls[k].name, stall_name):
            stall_name = canteen_name.stalls[k]
            break
    else:
        return False
    # this part is to update information to excel file
    for  i in range(1,sheet.max_row):
        if sheet.cell(row = i, column = 3).value == canteen_name.stalls[k].name: 
            for w in range(i, sheet.max_row):
                if sheet.cell(row = w, column = 4).value == str(food_name):
                    sheet.cell(row = w, column = 5).value = float(new_price)
                    sheet.cell(row = w, column = 8).value = int(new_rank)
                    break
                elif sheet.cell(row = w, column = 4).value == None:
                    sheet.insert_rows(i+1)
                    sheet.cell(row = i+1, column = 4).value = str(food_name)
                    sheet.cell(row = i+1, column = 5).value = float(new_price)
                    sheet.cell(row = i+1, column = 8).value = int(new_rank)
                    break
            break
    book.save('Canteen-Copy.xlsx') # save updated information to the screen
    book.close # close the excel file
    reload_data(canteen_name.name, stall_name.name) # reload the data to the program 
    return True


def type_of_stall():
    global gameDisplay, control, display_height, display_width,type_stall_name, type_of_food_image
    crashed = False
    # text font
    text_display = pygame.font.SysFont("Candara", 20)
    text_display_1 = pygame.font.SysFont("Calibri", 50)
    text_display_2 = pygame.font.SysFont("Calibri", 25)
    text_display_3 = pygame.font.SysFont("Book Antiqua", 22)
    # load image of star to used for get ranking input
    star_tranparent = pygame.image.load("star_transparent.jpg")
    star_color = pygame.image.load("star.jpg")
    star_tranparent = pygame.transform.smoothscale(star_tranparent, (40,40))
    star_color = pygame.transform.smoothscale(star_color, (40,40))
    # control variable
    number_of_star = 0
    blit_star = 0
    start_to_ask =  0 
    string_min = string_max = [] 
    start_to_blit = 0
    stall_name_result = ""
    check_price = 0 
    while not crashed:
        gameDisplay.fill((255,255,255)) # fill the screen with white color
        mouse_x = get_mouse_position()[0]
        mouse_y = get_mouse_position()[1]
        for event in pygame.event.get():
            if event.type == pygame.VIDEORESIZE: # resize window
                display_height = event.h
                display_width = event.w
                gameDisplay = pygame.display.set_mode((display_width,display_height),pygame.RESIZABLE)
            
            elif event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
                image_mouse_x = (mouse_x-40)//245
                if 5 < mouse_x <205 and 5 < mouse_y <35: # back button
                    crashed = True # back button
                    control = 1
                elif 185 < mouse_x < 265 and 435 < mouse_y < 465:
                    start_to_ask = 1 # get user input for min price
                elif 355 < mouse_x < 435 and 435< mouse_y< 465:
                    start_to_ask = 2  # get user output for max price
                elif 0 <= (mouse_x-600)/40 <= 5 and 450 < mouse_y < 500:
                    blit_star = 1 # get the number of stars used for ranking
                    number_of_star = (mouse_x-600)//40 
                elif 40 + 245*image_mouse_x < mouse_x < 235 +245*image_mouse_x and 75 < mouse_y< 230:
                    start_to_blit = 1 # change color of choosen type of cuisine
                    picture_chosen = image_mouse_x
                elif 40 + 245*image_mouse_x < mouse_x < 235 +245*image_mouse_x and 255 < mouse_y < 410:
                    start_to_blit = 2 
                    picture_chosen = image_mouse_x
                elif 800 < mouse_x < 980 and 10 < mouse_y < 60 and stall_name_result != "" and check_price == 2 :
                    get_result_of_stall_name(stall_name_result, string_min, string_max, number_of_star )
                else:
                    start_to_ask = 0  

            elif event.type == pygame.QUIT:
                pygame.quit()
                quit()
        text_3 = text_display_1.render("Choose your cuisine", 1, (0,50,50))
        gameDisplay.blit(text_3, (300,20))
        gameDisplay.blit(text_display_2.render("Price Min", 1, (0,0,100)),(80, 440))
        gameDisplay.blit(text_display_2.render("Max", 1, (0,0,100)),(300, 440))
        gameDisplay.blit(text_display_2.render("Rank", 1, (0,0,100)),(530, 440))
        pygame.draw.rect(gameDisplay,(0,0,0), (185, 438, 80,30), 1)
        pygame.draw.rect(gameDisplay,(0,0,0), (355, 438, 80,30), 1)
        # this part is to get and display input to the screen
        if start_to_ask == 1: # get user input for min price
            ask(event, string_min)
            string_min = string_min[:4]
        elif start_to_ask == 2: # get user input for max price
            ask(event, string_max)
            string_max = string_max[:4]
        
        #this is to check whether the input value of user for price is true or false
        try:
            float("".join(string_min))
            check_price = 2
        except Exception:
            if  len(string_min) != 0:
                check_price = 1
                gameDisplay.blit(text_display_3.render("Invalid Value", 1, (150,0,0)), (210, 460 ))
        try:
            float("".join(string_max))
            check_price = 2
        except Exception:
            if  len(string_max) != 0:
                check_price = 1
                gameDisplay.blit(text_display_3.render("Invalid Value", 1, (150,0,0)), (210, 472 ))
        gameDisplay.blit(text_display_3.render("".join(string_min), 1, (0,0,0)), (188, 438 ))
        gameDisplay.blit(text_display_3.render("".join(string_max), 1, (0,0,0)), (360, 438 ))
        
        # this part is to display text and draw rect of type stall
        for i in range(4):
            pygame.draw.line(gameDisplay, (0,0,0), (40 + 245*i, 195), (240 + 245*i, 195))
            pygame.draw.line(gameDisplay, (0,0,0), (40 + 245*i, 375), (240 + 245*i, 375))
            pygame.draw.rect(gameDisplay,(0,0,0), (40 + 245*i, 80, 200,150), 1)
            pygame.draw.rect(gameDisplay,(0,0,0), (40 + 245*i, 260, 200,150), 1)
            text_1 = text_display.render("%s"%(type_stall_name[::-1][i]),1, (0,0,0))
            text_2 = text_display.render("%s"%(type_stall_name[::-1][i+4]),1, (0,0,0))
            text_rect_1 = text_1.get_rect(center = (140 + 245*i, 212.5))
            text_rect_2 = text_2.get_rect(center = (140 + 245*i, 392.5) )
            gameDisplay.blit(text_1, text_rect_1)
            gameDisplay.blit(text_2, text_rect_2)
            gameDisplay.blit(type_of_food_image[i],(41 + 245*i, 81))
            gameDisplay.blit(type_of_food_image[i+4],(41 + 245*i, 261))
        # this part is to change the color of star rank if user choose a rank
        if blit_star == 1:
            for k in range(number_of_star + 1):
                gameDisplay.blit(star_color, (600 + 40*k, 430))
            for j in range( 4- number_of_star):
                gameDisplay.blit(star_tranparent, (760 - 40*j, 430))
        else:
            for j in range(5):
                gameDisplay.blit(star_tranparent, (600 + 40*j, 430))
        # change the color of the chosen cuisine
        if start_to_blit == 1:
            text_1 = text_display.render("%s"%(type_stall_name[::-1][picture_chosen]),1, (50,200,200))
            text_rect_1 = text_1.get_rect(center = (140 + 245*(picture_chosen), 212.5))
            gameDisplay.blit(text_1, text_rect_1)
            stall_name_result = type_stall_name[::-1][picture_chosen]
        elif start_to_blit ==2:
            text_2 = text_display.render("%s"%(type_stall_name[::-1][picture_chosen+4]),1, (50,200,200))
            text_rect_2 = text_2.get_rect(center = (140 + 245*picture_chosen, 392.5) )
            gameDisplay.blit(text_2, text_rect_2)
            stall_name_result = type_stall_name[::-1][picture_chosen + 4]
        # continue button
        pygame.draw.rect(gameDisplay, (0,0,0), (800, 10, 180, 50),1)
        gameDisplay.blit(text_display_2.render("Continue", 1, (0,0,100)),(820, 20))
        back()
        pygame.display.update()
        clock.tick(60)

def popoup_info(detail, order_of_canteen,delta_x, delta_y): # this function is used for popup if user's mouse is pointing towards a canteen
    global gameDisplay
    text_display_4 = pygame.font.SysFont("Candara", 23)
    text_display_5 = pygame.font.SysFont("Book Antiqua", 30)
    popup_up = pygame.image.load("popupbox.png")
    popup_down = pygame.image.load("popupinverse.png")
    popup_up = pygame.transform.smoothscale(popup_up, (600,300))
    popup_down = pygame.transform.smoothscale(popup_down, (600,300))
    if order_of_canteen< 6:
        gameDisplay.blit(popup_up, (300 +delta_x, 28+ order_of_canteen*390/12))
        gameDisplay.blit(text_display_5.render("%s"%(detail.name),1,(0,0,0)), (430+delta_x, 40+ order_of_canteen*390/12))
        gameDisplay.blit(text_display_4.render("Address : %s"%(detail.real_location),1,(0,0,0)), (430+delta_x, 100+ order_of_canteen*390/12))
        gameDisplay.blit(text_display_4.render("Seat Capacity %s"%(detail.seat_capacity),1,(0,0,0)), (430+delta_x, 140+ order_of_canteen*390/12))
        gameDisplay.blit(text_display_4.render("Open Time: %s"%(detail.time_open),1,(0,0,0)), (430+delta_x, 180+ order_of_canteen*390/12))
        gameDisplay.blit(text_display_4.render("Phone number: %s"%(detail.phone),1,(0,0,0)), (430+delta_x, 220+ order_of_canteen*390/12))
    else:
        gameDisplay.blit(popup_down, (300 +delta_x, delta_y -110+ order_of_canteen*390/12))
        gameDisplay.blit(text_display_5.render("%s"%(detail.name),1,(0,0,0)), (430+delta_x, -60+ order_of_canteen*390/12))
        gameDisplay.blit(text_display_4.render("Address : %s"%(detail.real_location),1,(0,0,0)), (430+delta_x, -20+ order_of_canteen*390/12))
        gameDisplay.blit(text_display_4.render("Seat Capacity %s"%(detail.seat_capacity),1,(0,0,0)), (430+delta_x,20+ order_of_canteen*390/12))
        gameDisplay.blit(text_display_4.render("Open Time: %s"%(detail.time_open),1,(0,0,0)), (430+delta_x, 60+ order_of_canteen*390/12))
        gameDisplay.blit(text_display_4.render("Phone number: %s"%(detail.phone),1,(0,0,0)), (430+delta_x,  100+ order_of_canteen*390/12))




def get_result_of_stall_name(stall_result_name, string_min, string_max, number_of_star): # this function is executed after user choose their type of cuisine and price range
    global gameDisplay, mergesort, list_Canteen, geo_location_y,geo_location_x, list_canteen_sortedbydistance
    crashed = False
    # load image to display
    popup_up = pygame.image.load("popupbox.png")
    popup_down = pygame.image.load("popupinverse.png")
    popup_up = pygame.transform.smoothscale(popup_up, (600,300))
    popup_down = pygame.transform.smoothscale(popup_down, (600,300))
    # get min, max value
    min_value = float("".join(string_min))
    max_value = float("".join(string_max))
    # searching for stall with given type of cuisine
    result = search_for_stall_name(stall_result_name, min_value, max_value, number_of_star)
    # sort the canteen by name
    canteen_name = mergesort(list(set([i[0].name for i in result])))
    # text font to display
    text_display = pygame.font.SysFont("Calibri", 28)
    text_display_1 = pygame.font.SysFont("Candara", 18)
    text_display_2 = pygame.font.SysFont("Book Antiqua", 20)
    text_display_3 = pygame.font.SysFont("Candara", 15)
    text_display_4 = pygame.font.SysFont("Candara", 23)
    text_display_5 = pygame.font.SysFont("Book Antiqua", 30)
    # control variables
    sort_distance = sort_price = sort_name = 0
    information = 0
    no_page = 0
    no_page_color = 1
    order_of_canteen= 0
    detail= 0 
    start_to_pop = 0
    get_direction = None
    canteen_direction = None
    
    while not crashed:
        gameDisplay.fill((255,255,255)) # fill the screen with white color
        # get the position of the mouse
        mouse_x = get_mouse_position()[0]
        mouse_y = get_mouse_position()[1]

        for event in pygame.event.get():
            if event.type == pygame.VIDEORESIZE: # resize window
                display_height = event.h
                display_width = event.w
                gameDisplay = pygame.display.set_mode((display_width,display_height),pygame.RESIZABLE)
            
            elif event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
                if 5 < mouse_x <205 and 5 < mouse_y <35: # back button
                    crashed = True
                if 250 < mouse_x < 280 and 58 < mouse_y < 88:
                    sort_distance = 1 # if user choose sort canteen by distance
                    sleep(0.1)
                elif 760 < mouse_x < 790 and 58 < mouse_y < 88:
                    sort_name = 1 # if suer choose to sort result of food by name
                    sort_price = 0
                elif 850 < mouse_x < 880 and 58 < mouse_y < 88:
                    sort_price = 1 # if user choose to sort result of food by price
                    sort_name = 0
                elif 21 < mouse_x < 299 and 92 < mouse_y < 482:
                    index_canteen = (mouse_y-92)*12//390 # get the name of chosen canteen
                    information = 1
                elif information == 1 and 420 <= mouse_x <= 420 + 20*no_page and 65 < mouse_y < 80:
                    no_page_color = (mouse_x - 400)//20 # fill the chosen number page with color
            elif event.type == pygame.QUIT:
                pygame.quit()
                quit()
        # this part is to display the name of canteen available with given criteria
        if sort_distance%2 !=0: # if user want to sort canteen by distance
            sorted_can = []
            for i in list_canteen_sortedbydistance:
                if i.name in canteen_name:
                    sorted_can.append(i.name)
            canteen_name = [i for i in sorted_can]
        no_canteen = len(canteen_name)
        for i in range(no_canteen): # display the name of canteen to screen
            gameDisplay.blit(text_display_2.render("%s"%(canteen_name[i]),1,(0,0,0)),(30,96 + 390/12*(i)))
            pygame.draw.line(gameDisplay, (0,0,0), (21,92 + 390/12*(i+1) ),(299,92 + 390/12*(i+1)),1)
        # display the text to screen
        gameDisplay.blit(text_display.render("%s Cuisine, Price: %s - %s, Rank: %s"%(stall_result_name, min_value, max_value, number_of_star+1), 1, (100,0,150)), (230,10))
        pygame.draw.line(gameDisplay,(0,0,0),(310,92),(310,482),1)
        gameDisplay.blit(text_display_1.render("Distance", 1, (0,0,100)), (170, 70))
        pygame.draw.rect(gameDisplay, (0,0,0), (250,58, 30, 30), 1)
        gameDisplay.blit(text_display_1.render("Sort: Name", 1, (0,0,100)), (650, 70))
        gameDisplay.blit(text_display_1.render("Price", 1, (0,0,100)), (800, 70))
        pygame.draw.rect(gameDisplay, (0,0,0), (760,58, 30, 30), 1)
        pygame.draw.rect(gameDisplay, (0,0,0), (850,58, 30, 30), 1)

        if sort_distance%2 == 1: # if user choosen sort canteen by distance box
            pygame.draw.rect(gameDisplay, (0,0,50), (250,58, 30, 30), 0)
        
        if sort_name == 1: # if user chooses to sort food by name
            pygame.draw.rect(gameDisplay, (0,0,50), (760,58, 30, 30), 0)
        
        if sort_price == 1:# if user choose to sort food by price
            pygame.draw.rect(gameDisplay, (0,0,50), (850,58, 30, 30), 0)
        
        if information == 1:
            gameDisplay.blit(text_display_2.render("Stall", 1, (0,50,100)), (370, 95))
            gameDisplay.blit(text_display_2.render("Food", 1, (0,50,100)), (625, 95))
            gameDisplay.blit(text_display_2.render("Price", 1, (0,50,100)), (905, 95))
            text_to_display = []
            # get the result of chosen canteen 
            for i in result:
                try:
                    if i[0].name == canteen_name[index_canteen]:
                        canteen_direction = i[0]
                        text_to_display.append((i[1],i[2],i[3]))
                except Exception:
                    pass
            if sort_name == 1 : #sort food by name  # use bubble sort
                for i in range(len(text_to_display)):
                    swap = 0
                    for k in range(len(text_to_display)- 1- i):
                        if text_to_display[k][1] > text_to_display[k+1][1]:
                            swap = 1
                            text_to_display[k], text_to_display[k+1] = text_to_display[k+1], text_to_display[k]
                    if swap == 0 :
                        break
            elif sort_price == 1 : #sort food by price  # use bubble sort
                for i in range(len(text_to_display)):
                    swap = 0
                    for k in range(len(text_to_display)- 1- i):
                        swap = 1
                        if text_to_display[k][2] > text_to_display[k+1][2]:
                            text_to_display[k], text_to_display[k+1] = text_to_display[k+1], text_to_display[k]
                    if swap == 0:
                        break
            # determine which page that user is choosing
            no_page = len(text_to_display)//15 +1
            for i in range(0,no_page):
                gameDisplay.blit(text_display_2.render("%s"%(i+1),1,(0,0,0)), (400 + 20*(i+1), 65))
            if no_page_color != 0 :
                gameDisplay.blit(text_display_2.render("%s"%(no_page_color),1,(100,100,210)), (400 + 20*(no_page_color), 65))
            
            for i in range(14): # display the information of food to the screen
                try: 
                    string = text_to_display[i+(no_page_color-1)*14]
                    gameDisplay.blit(text_display_3.render("%s"%(string[0]), 1, (0,0,0)),(325,100+ 390/15*(i+1)))
                    gameDisplay.blit(text_display_3.render("%s"%(string[1]), 1, (0,0,0)),(575,100+ 390/15*(i+1)))
                    gameDisplay.blit(text_display_3.render("%s"%(string[2]), 1, (0,0,0)),(915,100+ 390/15*(i+1)))
                except IndexError:
                    continue
            
            for i in range(1,15): # 
                pygame.draw.line(gameDisplay, (0,0,0), (325,92 + 390/15*(i+1) ),(930,92 + 390/15*(i+1)),1)
        
        if 21 < mouse_x < 299 and 92 < mouse_y < 482:# determine whether mouse position is pointing to a particular canteen to pop up
            start_to_pop = 1 
            order_of_canteen= (mouse_y-92)*12//390
        else:
            start_to_pop = 0
        
        if start_to_pop == 1: # pop up to give more information about chosen canteen
            try:
                for i in list_Canteen:
                    if i.name == canteen_name[order_of_canteen]:
                        detail  = i
                        break
                popoup_info(detail, order_of_canteen, 0,0)
            except Exception:
                pass
        
        try: # if user choose get direction
            get_direction = button_get_direction(event, mouse_x, mouse_y)
        except Exception:
            pass
        
        if get_direction == True  and canteen_direction != None: # execute get direction function if user chooose get direction
            pop_up_direction(geo_location_y, geo_location_x, canteen_direction)

        back() # back button
        pygame.display.update() # update to display
        clock.tick(60) # set fps


def direction_google_map(origin_y, origin_x, canteen_destination, mode): # this function is executed to get direction from google map
    global gameDisplay
    destination_y = canteen_destination.location_on_map[0]
    destination_x = canteen_destination.location_on_map[1]
    if mode == "bus": # if user choose to travel by bus
        url = "https://maps.googleapis.com/maps/api/directions/json?origin="+ str(origin_y)+","+str(origin_x)+"&destination="+str(destination_y)+","+str(destination_x)+"&mode=transit&transit_mode=bus&optimize_waypoints=True&key=AIzaSyC7YmRvn4WuGbFEG1Yy5DrQ3RM2VtuNXtU"
        r = requests.get(url)
        steps = r.json()["routes"][0]["legs"][0]["steps"]
        start_address = r.json()["routes"][0]["legs"][0]["start_address"]
        distance = r.json()["routes"][0]["legs"][0]["distance"]["text"]
        duration = r.json()["routes"][0]["legs"][0]["duration"]["text"]
        end_address = r.json()["routes"][0]["legs"][0]["end_address"]
        direction_detail = []
        for i in steps:
            if i["travel_mode"] != "WALKING":
                string = ""
                parser = MyHTMLParser()
                parser.feed(i["html_instructions"])
                string  += "".join(parser.return_data())
                string += " .Start at bus stop" +str(i["transit_details"]["departure_stop"]["name"])
                string += ". Take bus number " +str(i["transit_details"]["line"]["short_name"])
                string += ", pass " +str(i["transit_details"]["num_stops"]) +" bustops"
                string += ". Stop at "+str(i["transit_details"]["arrival_stop"]["name"])
                direction_detail.append(string)
            else:
                string = ""
                parser = MyHTMLParser()
                parser.feed(i["html_instructions"])
                string  += "".join(parser.return_data())
                string += "("+i["distance"]["text"]+" ,"+ i["duration"]["text"]+")"
                direction_detail.append(string)
    else: # if user want to travel by driving or walking
        url = "https://maps.googleapis.com/maps/api/directions/json?origin=" +str(origin_y)+","+str(origin_x)+"&destination="+str(destination_y)+","+str(destination_x)+"&mode="+str(mode)+"&optimize_waypoints=True&key=AIzaSyC7YmRvn4WuGbFEG1Yy5DrQ3RM2VtuNXtU"
        r = requests.get(url)
        steps = r.json()["routes"][0]["legs"][0]["steps"]
        start_address = r.json()["routes"][0]["legs"][0]["start_address"]
        distance = r.json()["routes"][0]["legs"][0]["distance"]["text"]
        duration = r.json()["routes"][0]["legs"][0]["duration"]["text"]
        end_address = r.json()["routes"][0]["legs"][0]["end_address"]
        direction_detail = []
        for i in steps:
            string = ""
            parser = MyHTMLParser()
            parser.feed(i["html_instructions"])
            string  += "".join(parser.return_data())
            string += "( about:"+i["distance"]["text"] + " ," +i["duration"]["text"]+")"
            direction_detail.append(string)            
    return start_address, end_address, duration, direction_detail, distance, url


def button_get_direction(event,mouse_x, mouse_y): #distance button
    global gameDisplay
    pygame.draw.rect(gameDisplay,(0,0,0),(840,10,150,27),1)
    information_text = pygame.font.SysFont("Arial", 20)
    gameDisplay.blit(information_text.render("Get direction",1,(0,0,0)),(847,13))
    if 840 < mouse_x < 990 and 10< mouse_y < 37:
        if event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
            return True
        else:
            gameDisplay.blit(information_text.render("Get direction",1,(100,100,0)),(847,13))
# load cover image
cover = pygame.image.load("modern.jpg")
cover = pygame.transform.smoothscale(cover, (2880, 1800))

def pop_up_direction(origin_y, origin_x, destination): # this function is executed when user wants to get direction to the canteen
    global gameDisplay, display_width, display_height, cover
    crashed = False
    # text list to display as button on the screen
    text_list = ["Driving", "Walking", "Bus"]
    guide_direction = [] # this list contains directions got by 3 above methods
    for i in text_list:
        guide_direction.append(direction_google_map(origin_y, origin_x, destination,i.lower() ))
    #text font to display
    text_display = pygame.font.SysFont("Candara", 40)
    text_display_1 = pygame.font.SysFont("Book Antiqua", 25)
    text_display_2 = pygame.font.SysFont("Book Antiqua",20)
    text_display_3 = pygame.font.SysFont("Calibri",20)
    # control variable
    travel_mode = ""
    mouse_x_choose = float("inf")
    mouse_y_choose = float("inf")
    result = None
    travel_mode_index = ""

    while not crashed:
        mouse_x = get_mouse_position()[0]
        mouse_y = get_mouse_position()[1]
        #gameDisplay.fill((255,255,255))
        for event in pygame.event.get():
            if event.type == pygame.VIDEORESIZE: # resize window
                display_height = event.h
                display_width = event.w
                gameDisplay = pygame.display.set_mode((display_width,display_height),pygame.RESIZABLE)
            
            elif event.type == pygame.MOUSEBUTTONDOWN and event.button == 1:
                mouse_x_choose = mouse_x # get the position of clicked mouse button
                mouse_y_choose = mouse_y # get the position of clicked mouse button
                if 5 < mouse_x <205 and 5 < mouse_y <35: # back button
                    crashed = True
            elif event.type == pygame.QUIT: # quit the program
                pygame.quit()
                quit()

        gameDisplay.blit(cover,(0,0)) # display the cover
        # display text to screen
        title = text_display.render("Get direction with Google Map",1, (150,200,255))
        title_rect = title.get_rect(center = (display_width/2,50))
        gameDisplay.blit(title, title_rect)
        # load frame of button to screen
        frame = pygame.image.load("popupdirection.jpg")
        frame = pygame.transform.smoothscale(frame, (int(display_width*0.75), int(display_height*0.85)))
        gameDisplay.blit(frame, (display_width*0.25, display_height*0.14))
        button = pygame.image.load("buttondirection.png")
        button = pygame.transform.smoothscale(button, (int(display_width*0.2), int(display_height*0.15)))
        # display three buttons
        for i in range(3):
            gameDisplay.blit(button, (int(display_width*0.05),int(display_height*(0.4+0.2*i))))
        
        # display text of button to the screen
        for i in range(3):
            if  int(display_width*0.05) <mouse_x_choose <int(display_width*0.05)+int(display_width*0.2) and \
            int(display_height*(0.4+0.2*i)) < mouse_y_choose < int(display_height*(0.4+0.2*i)) + int(display_height*0.15):
                text = text_display_1.render(str(text_list[i]),1,(200,150,50))
                travel_mode = text_list[i]
                travel_mode_index = i
            else:
                text = text_display_1.render(str(text_list[i]),1,(0,0,50))
            text_rect = text.get_rect(center = (int(display_width*0.15),int(display_height*(0.475+0.2*i))))
            gameDisplay.blit(text,text_rect)
        
        if travel_mode != "": # if user choose a travel mode
            result  = guide_direction[travel_mode_index]
            gameDisplay.blit(text_display_2.render(" Your startpoint: %s"%(result[0]), 1,(0,0,100)),(display_width*0.25 + 32,display_height*0.14 + 20))
            gameDisplay.blit(text_display_2.render(" Your endpoint: %s"%(result[1]), 1,(0,0,100)),(display_width*0.25 + 30,display_height*0.14 + 45))
            gameDisplay.blit(text_display_2.render(" Estimated time:"+ str(result[2])  + " "*10 +"Transport mode: "+str(travel_mode)+" "*10 +"Distance: "+str(result[4]), 1,(0,0,100)),(display_width*0.25 + 30,display_height*0.14 + 70))
            
            t = 0  # control variable
            # get data to display
            data_display = result[3]
            for j in range(len(data_display)):
                length = len(data_display[j])
                for i  in range(int(length/80)+1):
                    try:
                        string_1 = data_display[j][80*i:80*i+80]
                        gameDisplay.blit(text_display_3.render(" %s"%(string_1), 1,(0,0,100)),(display_width*0.25 + 30,display_height*0.14 + 23*t + 100))
                        t += 1
                        length -= 70
                    except Exception:
                        break
                if display_height*0.14 + 23*t + 100 > display_height:
                    break
        back() # back button
        pygame.display.update() # update information to the screen
        clock.tick(60) # set fps


quit_variable = False
distance_program =0
control = -1 #control variable of the flow of program

while not quit_variable:
    if control == -1:
        cover_page() # showing cover page of program
    elif control == 0:
        gameDisplay = pygame.display.set_mode((display_width,display_height),pygame.RESIZABLE)
        start_program() # start program 
    elif control == 1:
        ask_criteria()
    elif control == 2  :
        distance_criteria()
    elif control == 3:
        search_food_by_enter_name()
    elif control == 5:
        update_information()
    elif control ==6 :
        type_of_stall()
pygame.quit()
quit()