
import pygame
import requests

import openpyxl
book = openpyxl.load_workbook('Canteen-Copy.xlsx', data_only=True) # get data from excel file

from mergesort import *

from html.parser import HTMLParser
#get data from each sheet
worksheet_A = book.get_sheet_by_name('Canteen_A')
worksheet_B = book.get_sheet_by_name('Canteen_B')
worksheet_1 = book.get_sheet_by_name('Canteen_1')
worksheet_2 = book.get_sheet_by_name('Canteen_2')
worksheet_4 = book.get_sheet_by_name('Canteen_4')
worksheet_9 = book.get_sheet_by_name('Canteen_9')
worksheet_11 = book.get_sheet_by_name('Canteen_11')
worksheet_13 = book.get_sheet_by_name('Canteen_13')
worksheet_14 = book.get_sheet_by_name('Canteen_14')
worksheet_16 = book.get_sheet_by_name('Canteen_16')
worksheet_NIE = book.get_sheet_by_name('Canteen_NIE')
worksheet_NorthHill = book.get_sheet_by_name('Canteen_NorthHill')

class Canteen():
    def __init__(self,name, location_on_map, real_location, seat_capacity, time_open, phone):
        self.name = name
        self.stalls = [] # list of Stall objects
        self.location_on_map = location_on_map
        self.real_location = real_location
        self.seat_capacity = seat_capacity
        self.time_open = time_open
        self.phone = phone

type_stall_name = ['Others', 'Western Cuisine', 'Mixed Rice', 'Indian Cuisine','Beverages & Desserts',\
                 'Chinese Cuisine', 'Vegetarian Food', 'Mala']

class Stall():
    def __init__(self, name, type_stall):
        self.name = name
        self.menu = []
        self.type_stall = type_stall


Canteen_A = Canteen("Canteen_A", (1.347016,103.680244),"North Spine Plaza, 76 Nanyang Drive", 1838, "7am-3pm", 64658588)
Canteen_B = Canteen("Canteen_B", (1.342447,103.682390)," South Spine, 50 Nanyang Avenue", 1050,"7am-3pm", 67900355)
Canteen_1 = Canteen("Canteen_1", (1.346571,103.686025),"Hall 1, 21 NanyangCircle", 310, "7am-9pm", 63343033)
Canteen_2 = Canteen("Canteen_2", (1.348347,103.685456), "Hall 2, 35 Students Walk", 446, "7am-9pm", 63343033)
Canteen_4 = Canteen("Canteen_4", (1.344263,103.685321), "10 Nanyang Drive NTU Hall 4", 303, "7am-9pm", 68998600)
Canteen_9 = Canteen("Canteen_9", (1.352256,103.685257), "Hall 9, 50 Nanyang Avenue", 293,"7am-9pm",96923456)
Canteen_11 = Canteen("Canteen_11", (1.354904,103.686476), "Hall 11, 21 Nanyang Avenue", 210,"7am-9pm", 97866726)
Canteen_13 = Canteen("Canteen_13", (1.351716,103.681076), "Hall 13, 32 Nanyang Cresent", 210,"7am-9pm", 98510908)
Canteen_14 = Canteen("Canteen_14", (1.352718,103.682166), "Hall 14, 34 Nanyang Cresent", 270, "7am-9pm", 81127239)
Canteen_16 = Canteen("Canteen_16", (1.350296,103.680923), "Hall 16, 50 Nanyang Walk", 304, "7am-9pm",94505893)
Canteen_NIE = Canteen("Canteen_NIE", (1.348749,103.677622),"1 Nanyang Walk", 405, "7am-9pm", 67903888)
Canteen_NorthHill = Canteen("Canteen_NorthHill", (1.354422,103.688176), "NorthHill, 60 Nanyang Cresent", 440,"7am-9pm", 85080232)

list_worksheet = [worksheet_1, worksheet_11, worksheet_13, worksheet_14, worksheet_16, worksheet_2, worksheet_4,\
                 worksheet_9, worksheet_A, worksheet_B, worksheet_NIE, worksheet_NorthHill]

list_Canteen = [Canteen_1, Canteen_11, Canteen_13, Canteen_14, Canteen_16, Canteen_2, Canteen_4,\
                 Canteen_9, Canteen_A, Canteen_B, Canteen_NIE, Canteen_NorthHill]


for k in range(len(list_Canteen)):
    row = row_stall = 3
    try:
        while True:
            name_stall = list_worksheet[k].cell(row = row_stall, column = 3).value
            type_stall = list_worksheet[k].cell(row = row_stall, column = 2).value
            if name_stall != None:
                list_Canteen[k].stalls.append(Stall(name_stall, type_stall))
            row_stall += 1
            if row_stall == list_worksheet[k].max_row:
                break
    except IndexError:
        pass

    for i in list_Canteen[k].stalls:
        try:
            menu = []
            while list_worksheet[k].cell(row = row, column = 4).value != None:
                menu.append((list_worksheet[k].cell(row =row, column = 4).value, list_worksheet[k].cell(row = row,column =  5).value, list_worksheet[k].cell(row = row, column = 8).value))
                row += 1
            i.menu = menu
            row += 1
            if list_worksheet[k].max_row == row:
                break
        except IndexError:
            pass

def reload_data(canteen_name, stall_name):
    global list_worksheet, list_Canteen
    book = openpyxl.load_workbook('Canteen-Copy.xlsx', data_only=True)
    for i in range(len(list_worksheet)):
        if list_worksheet[i].title == canteen_name:
            list_worksheet[i] = book.get_sheet_by_name(str(canteen_name))
            row = 3
            for  k in list_Canteen[i].stalls:
                if k.name == stall_name:
                    try:
                        menu = []
                        while list_worksheet[i].cell(row = row, column = 3).value != stall_name:
                            row += 1
                        while list_worksheet[i].cell(row = row, column = 4).value != None:
                            menu.append((list_worksheet[i].cell(row =row, column = 4).value, list_worksheet[i].cell(row = row,column =  5).value, list_worksheet[i].cell(row = row, column = 8).value))
                            row += 1
                        k.menu = menu
                        row += 1
                    except Exception:
                        pass
                    break
            break



'''for i in range(len(list_Canteen)): #use bubblesort to sort the canteen by name
    swap = 0
    for k in range(len(list_Canteen)-1-i):
        if list_Canteen[k].name > list_Canteen[k+1].name:
            list_Canteen[k] ,list_Canteen[k+1] = list_Canteen[k+1], list_Canteen[k]
            list_worksheet[k], list_worksheet[k+1] = list_worksheet[k+1], list_worksheet[k]
            swap =1
    if swap == 0:
        break'''


for i in list_Canteen: #use bubblesort to sort the foodstalls by name
    for k in range(len(i.stalls)):
        swap = 0
        for j in range(len(i.stalls) - 1 - k):
            if i.stalls[j].name > i.stalls[j+1].name:
                i.stalls[j] ,i.stalls[j+1] = i.stalls[j+1], i.stalls[j]
                swap = 1
        if swap == 0:
            break

for i in list_Canteen: # use bubble sort to sort the dish in foodstall by name
    for k in i.stalls:
        for j in range(len(k.menu)):
            swap = 0
            for t in range(len(k.menu) - 1 - j):
                if k.menu[t][0] > k.menu[t+1][0]:
                    k.menu[t] ,k.menu[t+1] = k.menu[t+1], k.menu[t]
                    swap = 1
            if swap == 0:
                break

def search_food_by_name(food_name, price_sort, rank_sort): #find the food by searching name
    global list_Canteen
    information = []
    for i in list_Canteen:
        for k in i.stalls:
            for j in k.menu:
                if j[0] == food_name:
                    information.append((i,k.name, j[1], j[2]))
    # i.name is name of canteen, k.name is name of stall, j[1] is price of food, j[2] is rank of food
    if price_sort == 1:
        for i in range(len(information) - 1):
            swap = 0
            for k in range(len(information)-1-i):
                if information[k][2] > information[k+1][2]:
                    information[k], information[k+1] = information[k+1], information[k]
                    swap =1
    elif rank_sort == 1:
        for i in range(len(information) - 1):
            swap = 0
            for k in range(len(information)-1-i):
                if information[k][3] > information[k+1][3]:
                    information[k], information[k+1] = information[k+1], information[k]
                    swap =1
    return information

#this part is to search food by some initial charaters input from user
# this list contains all the name of food 
all_food_name = [] 

for i in list_Canteen:
    for k in i.stalls:
        for j in k.menu:
            if j[0] not in all_food_name:
                all_food_name.append(j[0])

all_food_name = mergesort(all_food_name)
def search_foodname_by_characters(characters):
    global all_food_name
    available_food = []
    charaters = characters.lower()
    t = k =0
    crashed = False
    while len(available_food) < 11 and not crashed:
        for i in range(t, len(all_food_name)):
            k = i
            if charaters == all_food_name[i].lower()[:len(characters)] and all_food_name[i] not in available_food:
                t = i + 1
                available_food.append(all_food_name[i])
                break

        if k == len(all_food_name) - 1:
            crashed = True
    t = 0
    crashed = False
    while len(available_food) < 11 and not crashed:
        for i in range(t, len(all_food_name)):
            k = i
            if charaters in all_food_name[i].lower() and all_food_name[i] not in available_food:
                t = i + 1
                available_food.append(all_food_name[i])
                break
        if k == len(all_food_name) - 1:
            crashed = True
    return available_food[:10]

#this function is used to search for food based on provided price, stall name and canteen name from user
def search_food_by_price(price_of_food, name_of_stall_to_search, number_of_star):
    # price_of_food is a tuple containing the maximum and minimun value of food price
    global list_Canteen 
    list_of_food_in_stall = name_of_stall_to_search.menu # this list contains the menu of stall 
    
    food_satisfy_price = []
    if number_of_star >0:
        for i in list_of_food_in_stall:
            if price_of_food[0] < float(i[1]) < price_of_food[1] and number_of_star == i[2]:
                food_satisfy_price.append(i)
    else:
       for i in list_of_food_in_stall:
            try:
                if price_of_food[0] < float(i[1]) < price_of_food[1]:
                    food_satisfy_price.append(i)
            except Exception:
                continue
    return food_satisfy_price




def search_stalls_by_name(stall_name):
    global list_Canteen
    canteen_have_this_stall = []
    for i in list_Canteen:
        for k in i.stalls:
            if k.name == stall_name:
                canteen_have_this_stall.append(i.name)
                break
    return canteen_have_this_stall





def search_by_type_stall(type_stall_input):
    global location_type_stall, list_Canteen
    for i in location_type_stall:
        if i.name == type_stall_input:
            return i.location

def check_character(a,b):
    a = a.lower()
    letter_check = [chr(97 + i) for i in range(26)]+ [str(i) for i in range(9)]
    letter_a = [ i for i in a if i in letter_check]
    letter_b = [ i for i in b.lower() if i in letter_check]
    if letter_a == letter_b:
        return True
    return False


def search_for_stall_name(stall_name, min_value, max_value, rank):
    result =[]
    for i in list_Canteen:
        for k in i.stalls:
            if k.type_stall == stall_name:
                for j in k.menu:
                    try:
                        if min_value<= j[1] <= max_value and j[2] == rank:
                            result.append((i, k.name,j[0],j[1] ))
                    except Exception:
                        continue
    return result

type_of_food_image = []
for i in range(1,9):
    try:
        image = pygame.image.load("icon %s.png"%(i))
        image = pygame.transform.smoothscale(image, (198,114))
        type_of_food_image.append(image)
    except Exception:
        image = pygame.image.load("food.jpg")
        image = pygame.transform.scale(image, (198,114))
        type_of_food_image.append(image)

# this class is used to extract data from html instructions of google map 
class MyHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.my_data = []

    def handle_data(self, data):
        self.my_data.append(data)

    def return_data(self):
        return self.my_data



